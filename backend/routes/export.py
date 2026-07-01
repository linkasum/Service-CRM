"""
Маршруты: Экспорт в Excel
"""
import io
from datetime import datetime
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select, or_
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from core.database import get_session
from core.security import get_current_user
from models.user import User
from models.order import Order
from models.part import Part
from models.service import Service
from models.order_payment import OrderPayment, PaymentStatus, PaymentType
from core.logging import logger

router = APIRouter(prefix="/api/export", tags=["Экспорт"])


def _style_header(ws):
    """Стиль заголовка"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.freeze_panes = 'A2'


def _auto_width(ws):
    """Автоширина колонок"""
    for col in ws.columns:
        max_len = max((len(str(cell.value) or "") for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)


@router.get("/orders", summary="Экспорт заказов в Excel")
def export_orders(
    status: str = Query(None),
    date_from: str = Query(None),
    date_to: str = Query(None),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Экспорт заказов в Excel"""
    query = select(Order)
    if status:
        query = query.where(Order.status == status)
    if date_from:
        query = query.where(Order.created_at >= date_from)
    if date_to:
        query = query.where(Order.created_at <= date_to + " 23:59:59")

    orders = session.exec(query.order_by(Order.created_at.desc())).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"

    # Заголовки
    headers = [
        "ID", "Дата", "Клиент", "Телефон", "Email", "Устройство", "Бренд", "Модель",
        "Неисправность", "Статус", "Мастер", "Запчасти ₽", "Услуги ₽", "Итого ₽",
        "Оплачено ₽", "Остаток ₽", "Гарантия", "Дата выдачи"
    ]
    ws.append(headers)
    _style_header(ws)

    status_labels = {
        'new': 'Новый', 'diagnostics': 'Диагностика', 'agreed': 'Согласование',
        'repair': 'В работе', 'waiting_parts': 'Ждёт запчасти', 'ready': 'Готов',
        'issued': 'Выдан', 'cancelled': 'Отменён'
    }

    for o in orders:
        master_name = o.master.username if o.master else (o.acceptor.username if o.acceptor else "—")
        paid = session.exec(
            select(OrderPayment).where(
                OrderPayment.order_id == o.id,
                OrderPayment.status == PaymentStatus.completed,
                OrderPayment.payment_type != PaymentType.expense,
            )
        ).all()
        total_paid = sum(p.amount for p in paid)
        remaining = (o.total_cost or 0) - total_paid

        ws.append([
            o.id,
            o.created_at.strftime("%d.%m.%Y %H:%M") if o.created_at else "",
            o.client_name,
            o.client_phone,
            o.client_email or "",
            f"{o.device_brand} {o.device_model}",
            o.device_brand,
            o.device_model,
            (o.complaint or "")[:50],
            status_labels.get(o.status, o.status),
            master_name,
            o.parts_cost or 0,
            o.work_cost or 0,
            o.total_cost or 0,
            total_paid,
            remaining,
            f"{o.warranty_days or 0} дн." if o.warranty_days else "—",
            o.issued_at.strftime("%d.%m.%Y") if o.issued_at else "—",
        ])

    _auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"orders_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/parts", summary="Экспорт склада в Excel")
def export_parts(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Экспорт склада в Excel"""
    parts = session.exec(select(Part).order_by(Part.name)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Склад"

    headers = ["ID", "Название", "Артикул", "Кол-во", "Цена закупки", "Цена продажи", "Сумма склада", "Создан"]
    ws.append(headers)
    _style_header(ws)

    for p in parts:
        ws.append([
            p.id, p.name, p.article, p.quantity,
            p.cost_price, p.sale_price,
            p.sale_price * p.quantity,
            p.created_at.strftime("%d.%m.%Y") if p.created_at else "",
        ])

    _auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"parts_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/services", summary="Экспорт услуг в Excel")
def export_services(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Экспорт услуг в Excel"""
    services = session.exec(select(Service).order_by(Service.name)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Услуги"

    headers = ["ID", "Название", "Описание", "Цена", "Статус", "Создан"]
    ws.append(headers)
    _style_header(ws)

    for s in services:
        ws.append([
            s.id, s.name, (s.description or "")[:50],
            s.price, "Активна" if s.status.value == "active" else "Неактивна",
            s.created_at.strftime("%d.%m.%Y") if s.created_at else "",
        ])

    _auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"services_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/clients", summary="Экспорт клиентов в Excel")
def export_clients(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Экспорт клиентов в Excel"""
    # Группируем по телефону
    orders = session.exec(select(Order).order_by(Order.created_at)).all()
    clients = {}
    for o in orders:
        key = o.client_phone
        if key not in clients:
            clients[key] = {
                "name": o.client_name,
                "phone": o.client_phone,
                "email": o.client_email or "",
                "orders": 0,
                "total": 0,
                "last_order": "",
            }
        clients[key]["orders"] += 1
        clients[key]["total"] += o.total_cost or 0
        clients[key]["last_order"] = o.created_at.strftime("%d.%m.%Y") if o.created_at else ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"

    headers = ["Клиент", "Телефон", "Email", "Заказов", "Общая сумма", "Последний заказ"]
    ws.append(headers)
    _style_header(ws)

    for c in clients.values():
        ws.append([c["name"], c["phone"], c["email"], c["orders"], c["total"], c["last_order"]])

    _auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"clients_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
